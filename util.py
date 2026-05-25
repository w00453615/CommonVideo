import re
from datetime import datetime
from random import random, randint
from typing import Optional

from ascript.android.system import open
from ascript.android import action
# __init__.py 为初始入口文件,工程代码的入口文件.

# 导入动作库常用函数
from ascript.android.action import click, slide, Touch, gesture
# 导入控件检索相关
from ascript.android.node import Selector
# 导入图色相关
from ascript.android.screen import capture, FindColors, FindImages, Ocr
# 导入系统相关
from ascript.android import system
# 环境设备相关
from ascript.android.system import R, Device
from ascript.android.ui import Dialog
from .logger.yunpanlogger import logger

import math
import time

curApp = None
WAIT_LOW = 30
WAIT_HIGH = 35

def clickNode(node):
    logger.info(f'clickNode {node}')
    centerX = node.rect.centerX() + math.floor(random() * 11) - 5
    centerY = node.rect.centerY() + math.floor(random() * 11) - 5
    logger.info(f'clickNode {centerX} {centerY}')
    click(centerX, centerY)

def clickXY(x, y, delta = 5):
    centerX = randint(x-delta, x+delta)
    centerY = randint(y-delta, y+delta)
    # print(centerX, centerY)
    click(centerX, centerY)

def findDescAndClick(keyword, delay=5):
    logger.info(f'findDescAndClick {keyword}')
    for _ in range(delay):
        node = Selector(0|2).desc(keyword).find()
        if node:
            clickNode(node)

            return node
        else:
            wait_with_jitter(2)
    print(f'findDescAndClick exit {keyword}')
    return None


def findPathAndClick(keyword, delay=5):
    logger.info(f'findPathAndClick {keyword}')
    for _ in range(delay):
        node = Selector(0|2).path(keyword).find()
        if node:
            clickNode(node)
            return node
        else:
            wait_with_jitter(2)

    logger.info(f'findPathAndClick exit {keyword}')
    return None


def findTextAndClick(keyword, delay=5):
    logger.info(f'findTextAndClick {keyword}')
    for _ in range(delay):
        node = Selector(0|2).text(keyword).find()

        if node:
            clickNode(node)

            return node
        else:
            wait_with_jitter(2)
    logger.info(f'findTextAndClick exit {keyword}')
    return None

def findText(keyword, delay=5, index=0):
    logger.info(f'findText {keyword}')
    keyword = keyword.replace('\\\\', '\\')
    for _ in range(delay):
        if index == 0:
            node = Selector(0|2).text(keyword).find()
        else:
            all = Selector(0|2).text(keyword).find_all()
            # print(f'findTextAll {all}')
            if all is None:
                return None
            elif len(all) <= index:
                node = all[0]
            else:
                node = all[index]

        if node:
            logger.info(f'findText end {node.text}')
            return node.text
        else:
            wait_with_jitter(2)
    logger.info(f'findText exit {keyword}')
    return None

def findDesc(keyword, delay=5):
    logger.info(f'findDesc {keyword}')
    for _ in range(2):
        node = Selector(0|2).desc(keyword).find()

        if node:
            logger.info(f'findDesc found {node.desc}')
            return node.desc
        else:
            wait_with_jitter(2)
    logger.info(f'findDesc not found {keyword}')
    return None

def findImageAndClick(keyword, delay=5):
    logger.info(f'findImageAndClick {keyword}')
    for _ in range(delay):
        res = FindImages.find_template([R.img(keyword), ], confidence=0.7)

        if res:
            logger.info(f'findImageAndClick {res}')
            x = res["center_x"] + randint(-10, 20)
            y = res["center_y"] + randint(-10, 20)

            click(x, y)

            return True
        else:
            wait_with_jitter(2)
    logger.info(f'findImageAndClick exit {keyword}')
    return None


def findTextEndAndClick(keyword, end, delay=5):
    logger.info(f'findTextEndAndClick {keyword}')
    for _ in range(delay):
        node = Selector(0|2).text(keyword).find()

        if node:
            clickNode(node)
            logger.info(f'findTextEndAndClick1 end {node}')
            return node
        else:
            wait_with_jitter(2)
        logger.info(f'findTextEndAndClick1 end {node}')
        node2 = Selector(0|2).text(end).find()
        if node2:
            logger.info(f'findTextEndAndClick2 end {node2}')
            return node
    logger.info(f'findTextEndAndClick exit {keyword}')
    return None


def swipeBackOnes():
    d = Device.display()
    weight = d.widthPixels
    action.slide(weight - 2 - math.floor(random() * 5), 500 + math.floor(random() * 11) - 5,
                 weight - 260 + math.floor(random() * 11) - 5, 550 + math.floor(random() * 11) - 5,
                 300 + random() * 200)
def swipeBackCount(count):
    for _ in range(count):
        d = Device.display()
        weight = d.widthPixels
        action.slide(weight - 2 - math.floor(random() * 5), 500 + math.floor(random() * 11) - 5,
                     weight - 260 + math.floor(random() * 11) - 5, 550 + math.floor(random() * 11) - 5,
                     300 + random() * 200)
        time.sleep(0.2 + random() / 2)


# def swipeBack(to, to2=None):
#     print('swipeBack')
#
#     count = 0
#     time.sleep(2 + random() / 2)
#     while True:
#         x, y = ocrFind(to)
#         if x is not None:
#             if to2 is not None:
#                 x2, y2 = ocrFind(to2)
#                 if x2 is not None:
#                     break
#             else:
#                 break
#         action.slide(1060 + math.floor(random() * 11) - 5, 500 + math.floor(random() * 11) - 5,
#                      800 + math.floor(random() * 11) - 5, 550 + math.floor(random() * 11) - 5, 300 + random() * 200)
#         time.sleep(2 + random() / 2)
#         count += 1
#         if count > 3:
#             return False
#     return True

def swipeBack(to, type='ocr'):
    # print('swipeBack')
    x = None
    y = None
    count = 0
    time.sleep(1 + random() / 2)

    while True:
        if type == 'ocr':
            x, y = ocrFind(to)
        elif type == 'desc':
            x, y = descFind(to)
        elif type == 'text':
            x, y = textFind(to)
        if x is not None:
            break
        d = Device.display()
        weight = d.widthPixels
        action.slide(weight - 2 - math.floor(random() * 5), 500 + math.floor(random() * 11) - 5,
                     weight - 260 + math.floor(random() * 11) - 5, 550 + math.floor(random() * 11) - 5,
                     300 + random() * 200)

        time.sleep(2 + random() / 2)
        count += 1
        if count > 3:
            return False
    return True

def swipeBackApp(to, type='ocr'):
    if swipeBack(to, type) == False:
        print('swipeBackApp fail retry', curApp)
        wait_with_jitter(2)
        action.Key.home()
        wait_with_jitter(2)
        print('swipeBackApp open', curApp)
        system.open(curApp)
        swipeBack(to, type)

def swipeDownTo(to, type='ocr'):
    # print('swipeUp')
    x = None
    y = None
    count = 0
    time.sleep(1 + random() / 2)

    while True:
        if type == 'ocr':
            x, y = ocrFind(to)
        elif type == 'desc':
            x, y = descFind(to)
        elif type == 'text':
            x, y = textFind(to)
        if x is not None:
            break
        d = Device.display()
        weight = d.widthPixels
        swipeDown()

        time.sleep(1 + random() / 2)
        count += 1
        if count > 5:
            return False
    return True

def swipeUp(low=0.2, high=0.7):
    size = Device.display()
    width = int(size.widthPixels * 0.5)
    startX = randint(width - 10, width + 10)
    startY = randint(int(size.heightPixels * high) - 10, int(size.heightPixels * high) + 10)
    endX = randint(width + 10, width + 20)
    endY = randint(int(size.heightPixels * low) - 10, int(size.heightPixels * low) + 10)
    # action.slide(500 + randint(0, 100), 1800 + randint(0, 100),
    #              500 + randint(0, 100), 1100 + randint(0, 100), 200 + randint(100, 200))

    # action.slide(startX, startY, startX, startY - 100, randint(95, 105))
    # action.slide(startX, startY-100, startX, startY-300, randint(95, 105))
    # action.slide(startX, startY - 300, startX, endY, randint(95, 105))
    t = randint(200, 220)
    action.slide(startX, startY, endX, endY, t)
    print(f'swipeUp {startX}, {startY}, {endX}, {endY}, {t}')


def swipeDown(low=0.2, high=0.7):
    size = Device.display()
    width = int(size.widthPixels * 0.5)
    startX = randint(width - 10, width + 10)
    endY = randint(int(size.heightPixels * high) - 10, int(size.heightPixels * high) + 10)
    endX = randint(width + 10, width + 20)
    startY = randint(int(size.heightPixels * low) - 10, int(size.heightPixels * low) + 10)
    # action.slide(500 + randint(0, 100), 1800 + randint(0, 100),
    #              500 + randint(0, 100), 1100 + randint(0, 100), 200 + randint(100, 200))

    # action.slide(startX, startY, startX, startY - 100, randint(95, 105))
    # action.slide(startX, startY-100, startX, startY-300, randint(95, 105))
    # action.slide(startX, startY - 300, startX, endY, randint(95, 105))
    t = randint(200, 220)
    action.slide(startX, startY, endX, endY, t)
    print(f'swipeDown {startX}, {startY}, {endX}, {endY}, {t}')


def swipes(sec=15):
    left = sec
    for _ in range(sec):
        print('swipes left', left)
        if left <= 0:
            break
        swipeUp()
        time.sleep(2 + random())
        swipeDown()
        time.sleep(2 + random())
        left = left - 4

        if Selector(0|2).text('任务已完成').find():
            return


def textFind(text):
    text = text.replace('\\\\', '\\')
    node = Selector(0|2).text(text).find()
    if node:
        return node.rect.centerX(), node.rect.centerY()
    else:
        return None, None

def idFind(text):
    node = Selector(0|2).id(text).find()
    if node:
        return node.rect.centerX(), node.rect.centerY()
    else:
        return None, None

def descFind(text):
    text = text.replace('\\\\', '\\')
    node = Selector(0|2).desc(text).find()
    if node:
        return node.rect.centerX(), node.rect.centerY()
    else:
        return None, None

def pathFind(text):
    print('pathFind', text)
    node = Selector(0|2).path(text).find()
    if node:
        print('pathFind', text, node)
        return node.rect.centerX(), node.rect.centerY()
    else:
        print('pathFind not find', text)
        return None, None

def imageFind(text, confidence=0.9):
    res = FindImages.find_template([R.img(text), ], confidence=confidence)
    if res:
        return res["center_x"], res["center_y"]
    else:
        return None, None

def ocrFindText(text, confidence=0.9, rect=None):
    # print('ocrFind', text)
    res = Ocr.mlkitocr_v2(rect=rect, pattern=text)
    if res:
        for r in res:
            # print(f'ocrFindText {r.text} end')
            if text is not None:
                # print(r.text)  # 打印出文本
                return r.center_x, r.center_y, r.text
                # print(r['center_x'], r['center_y'])  # 识别范围
                # print(r['confidence'])  # 可信度

    # print('ocrFind not find', text)
    return None, None, None

def ocrFindTextPad(text, confidence=0.9):
    # print('ocrFind', text)
    res = Ocr.paddleocr(None, pattern=text)
    if res:
        for r in res:
            print(f'ocrFindText {r.text} end')
            if text is not None:
                # print(r.text)  # 打印出文本
                return r.center_x, r.center_y, r.text
                # print(r['center_x'], r['center_y'])  # 识别范围
                # print(r['confidence'])  # 可信度

    # print('ocrFind not find', text)
    return None, None, None

def ocrFindTextRect(text, rect=None, confidence=0.9):
    # print('ocrFind', text)
    res = Ocr.mlkitocr_v2(rect=rect, pattern=text)
    if res:
        for r in res:
            print('ocrFindText', r.text, 'end')
            if text is not None:
                # print(r.text)  # 打印出文本
                return r.center_x, r.center_y, r.rect
                # print(r['center_x'], r['center_y'])  # 识别范围
                # print(r['confidence'])  # 可信度

    # print('ocrFind not find', text)
    return None, None, None

def ocrFind(text, confidence=0.9):
    x, y, _ = ocrFindText(text, confidence)
    return x,y
def swipe(paras):
    delay = paras['delay']
    count = paras['count']
    for _ in range(count):
        swipeUp()
        time.sleep(delay + randint(0, 20))


def forWait(paras, node):
    delay = paras['delay']
    count = paras['count']
    for _ in range(count):
        time.sleep(delay + randint(0, 20))

def findWaitBack(text, back, type):
    x, y = ocrFind(text)
    if x is not None:
        clickXY(x, y)
        left = randint(WAIT_LOW, WAIT_HIGH)
        logger.info(f'time {left} sleep {text}')
        time1 = datetime.now()
        while True:
            swipeUp()
            time.sleep(4 + random())

            time2 = datetime.now()
            time_diff = (time2 - time1).total_seconds()
            # print('等待时间', time_diff)
            if time_diff > left:
                break
        swipeBackApp(back, type)
    return x, y

def extract_first_num_to_int(s: str) -> Optional[int]:
    """
    提取字符串中第一个出现的连续数字段并转为int，无数字返回None
    :param s: 待处理字符串（数字前可能有字母/符号）
    :return: 转换后的整数（成功）/None（失败）
    """
    # 正则匹配：\d+ 匹配1个及以上连续数字；search找任意位置的第一个匹配
    match = re.search(r'\d+', s)
    if not match:
        print(f"字符串 '{s}' 中未找到数字")
        return None

    # 提取数字字符串并转换为int（自动忽略前导0）
    try:
        num_str = match.group()
        return int(num_str)
    except ValueError:
        # 理论上\d+匹配的字符串不会触发此异常，仅做兜底
        logger.info(f"数字字符串 '{num_str}' 转换失败")
        return None
def lookGuangGao(paras, x, y):
    for _ in range(10):
        time.sleep(3 + randint(0, 2))
        if textFind('领取成功'):
            swipeBackOnes()
        else:
            time.sleep(5 + random())


def getJinBi(paras):
    # print('getJinBi')
    if 'jinBiPos' not in paras:
        # print('getJinBi exit')
        return 0
    jinBiPos = paras['jinBiPos']
    rect = jinBiPos()
    # print('getJinBi', jinBiPos, 'rect', rect)

    # res = Ocr(rect=rect).paddleocr_v3()
    res = Ocr.mlkitocr_v2(rect, r'^[1-9]\d*$')
    if res:
        for r in res:
            # print(r)
            logger.info(f'getJinBi {r.text}')  # 打印出文本
            return int(r.text)
            # print(r['center_x'], r['center_y'])  # 识别范围
            # print(r['confidence'])  # 可信度
        # print('end')


def lookShortVideo(paras, x, y, low = 4, up = 5):
    wait_with_jitter(3)

    before = getJinBi(paras)
    count = 0
    for _ in range(2):
        time.sleep(randint(low, up))
        after = getJinBi(paras)
        if after is None:
            # time.sleep(5)
            swipeUp(0.2,0.6)

            continue
        # print(f'lookShortVideo {before} {after} {count}')
        if before == after:
            count += 1
        else:
            before = after
            count = 0
        if count >= 2:
            swipeUp()
            before = getJinBi(paras)

def clickNodeP(paras, x, y):
    clickXY(x, y)


def swipeUpP(paras, node):
    delay = paras['delay']
    delay = randint(delay, delay + 10)
    for _ in range(0, delay, 3):
        time.sleep(randint(3, 5))
        swipeUp()


procList = {}


def register(cfg):
    # print('register', cfg)
    procList[cfg['name']] = cfg


def findPath(step, name):
    path = step['path']
    if len(path) == 0:
        return True
    for it in path:
        parts = it.split(':', 2)
        type = parts[0]
        value = parts[1]
        delay = float(parts[2]) if len(parts) > 2 else 2
        logger.info(f'findPath1 {it} name {name} type {type} value {value} delay {delay}')
        time.sleep(3+random())
        
        find = False
        for _ in range(4):
            x,y = findPosSingle(type, value)
            if x is not None:
                logger.info(f'findPath2 found {it} {name} x {x} y {y} find {find}')
                clickXY(x, y)
                find = True
                if delay > 0:
                    logger.info(f'findPath2 wait {delay}s before returning')
                    time.sleep(delay+random())
                break
            
            logger.info(f'findPath3 not found swipeUp {it} {name}')
            swipeUp(0.2, 0.6)
            time.sleep(2+random())
        logger.info(f'findPath4 {type} {value} find {find}')
        if not find:
            logger.info(f'findPath5 {type} {value} not found')
            # for _ in range(4):
            #     swipeDown(0.2, 0.7)
            #     time.sleep(0.5+random())
            return False
    return True


def wait_with_jitter(seconds):
    """等待指定时间，并在基础上增加0-1秒的随机延迟"""
    jitter = random()
    total_time = seconds + jitter
    logger.info(f'wait_with_jitter {total_time:.2f}s')
    time.sleep(total_time)


def save_step_screenshot(step_name):
    """保存当前 step 的截图并上传到阿里云盘"""
    try:
        screenshot = capture()
        if screenshot:
            filename = f"{step_name}.png"
            if hasattr(logger, 'device_name') and hasattr(logger, 'run_date') and hasattr(logger, 'cfg_start_time') and hasattr(logger, 'current_cfg_name'):
                folder_path = f"log/{logger.device_name}/{logger.run_date}/{logger.cfg_start_time}/{logger.current_cfg_name}"
                
                if hasattr(screenshot, 'compress'):
                    from android.graphics import Bitmap
                    from java.io import ByteArrayOutputStream
                    
                    stream = ByteArrayOutputStream()
                    screenshot.compress(Bitmap.CompressFormat.PNG, 100, stream)
                    stream.flush()
                    screenshot_data = stream.toByteArray()
                    stream.close()
                    
                    logger.upload_image(screenshot_data, filename, folder_path)
    except Exception as e:
        logger.info(f'保存step截图失败: {e}')


def procSingle(cfgName, check, run_timestamp=None, run_date=None):
    cfg = procList[cfgName]
    app = cfg['app']

    global curApp
    curApp = app

    system.open(app)
    wait_with_jitter(4)

    steps = cfg['step']
    nodes = cfg['node']
    name = cfg['name']
    logger.start_cfg(cfgName, Device.name(), run_timestamp, run_date)
    step_gold_data = {}
    
    for i in range(1):
        initial_gold = 0
        
        for step in steps:
            if i < 2 and step['name'] not in check:
                continue

            logger.start_step(step['name'])
            
            logger.info(f"procSingle log {cfgName} {step['name']}")
            
            if not findPath(step, step['name']):
                logger.info(f'procSingle {cfgName} {step["name"]} not found')
                final_gold = callStopCallbacks(nodes, step['name'])
                step_gold_data[step['name']] = max(0, final_gold - initial_gold)
                logger.flush(cfgName, step['name'])
                save_step_screenshot(step['name'])
                initial_gold = final_gold
                continue

            find = False

            for _ in range(1):
                nodeList = nodes[step['name']]
                logger.info(f'nodeList {nodeList}')
                
                for it in nodeList:
                    if 'process' not in it:
                        continue
                    proc = it['process']
                    logger.info(f'find node {it["name"]} proc {proc}')

                    if proc:
                        if 'time' in it:
                            procTime = it['time']
                        else:
                            procTime = 600
                        time1 = datetime.now()
                        while True:
                            time.sleep(2 + random())
                            if proc(it, None, None) == -1:
                                break
                            time2 = datetime.now()
                            time_diff = (time2 - time1).total_seconds()
                            print(f'procSingle {proc} time {time_diff}')
                            if time_diff > procTime:
                                break

                    else:
                        break
                final_gold = callStopCallbacks(nodes, step['name'])
                if initial_gold == 0:
                    initial_gold = final_gold
                step_gold_data[step['name']] = max(0, final_gold - initial_gold)
                logger.info(f'procSingle {cfgName} {step["name"]} gold {initial_gold} -> {final_gold}')
                initial_gold = final_gold
                logger.flush(cfgName, step['name'])
                save_step_screenshot(step['name'])
    
    return step_gold_data

                    
def callStopCallbacks(nodes, step_name):
    gold_coins = 0
    if step_name in nodes:
        nodeList = nodes[step_name]
        for it in nodeList:
            if 'stop' in it and callable(it['stop']):
                print(f'calling stop callback for {it["name"]}')
                result = it['stop']()
                if isinstance(result, (int, float)):
                    gold_coins = result
    return gold_coins

def findPosSingle(type, text):
    x = None
    y = None
    print(f'findPosSingle {type} {text}')
    if 'text' == type:
        x, y = textFind(text)
    elif 'desc' == type:
        x, y = descFind(text)
    elif 'path' == type:
        x, y = pathFind(text)
    elif 'id' == type:
        x, y = idFind(text)
    # elif 'image' == type:
    #     print(step['name'] + '/' + it['image'])
    #     x, y = imageFind(name + '/' + it['image'] + '.png')
    elif 'ocr' == type:
        text = text.replace('\\\\', '\\')
        x, y = ocrFind(r'' + text)
        # print(f'findPosSingle xy {x} {y}')
        if x is not None:
            return x, y
    elif 'rightocr' == type:
        text = text.replace('\\\\', '\\')
        resList = Ocr.mlkitocr_v2(pattern=r'' + text)
        if resList is not None:
            for res in resList:
                x, y = res.rect[2]-10, res.center_y
                print(f'findPosSingle rightocr {text} x {x} y {y} rect {res.rect} ')
                if x is not None:
                    return x, y
    elif 'padocr' == type:
        x, y, _ = ocrFindTextPad(r'' + text)
        print(f'findPosSingle rightocrPad {text} x {x} y {y} rect {res["rect"]} ')
        if x is not None:
            return x, y
    return x, y

def findPos(it, name, step):
    x = None
    y = None
    wait_with_jitter(3)
    print(f'findPos {it} step {step}')

    for _ in range(4):
        if 'text' in it:
            x, y = findPosSingle('text', it['text'])
        elif 'desc' in it:
            x, y = findPosSingle('desc',it['desc'])
        elif 'path' in it:
            x, y = findPosSingle('path',it['path'])
        elif 'id' in it:
            x, y = findPosSingle('path',it['id'])
        elif 'image' in it:
            print(f'findPos {step["name"]} {it["image"]}')
            x, y = imageFind(name + '/' + it['image'] + '.png')
        elif 'ocr' in it:
            for _ in range(2):
                text = it['ocr'].replace('\\\\','\\')
                x, y = findPosSingle('ocr', r''+text)
                print(f'findPos {text} {x} {y}')
                if x is not None:
                    return x, y

        else:
            return None, None
        if x is not None:
            break
        else:
            swipeUp(0.2, 0.5)
            wait_with_jitter(3)
    print(f'findPos {it["name"]} {x} {y}')
    return x, y


def read_existing_report_data(excel_buffer):
    """从现有Excel报表中读取数据，返回all_data字典"""
    import openpyxl
    
    if not excel_buffer:
        return {}
    
    try:
        wb = openpyxl.load_workbook(excel_buffer, data_only=True)
        ws = wb.active
        
        all_data = {}
        current_cfg = None
        
        row = 5  # 从第5行开始读取数据
        while row <= ws.max_row:
            cell_value = ws.cell(row=row, column=1).value
            
            # 如果是合并单元格，并且字体是粗体，说明是cfg名称
            if cell_value and ws.cell(row=row, column=1).font.bold:
                current_cfg = cell_value
                all_data[current_cfg] = {}
                row += 2  # 跳过表头
            elif current_cfg and cell_value and cell_value != '合计':
                step_name = cell_value
                gold_value = ws.cell(row=row, column=2).value
                if gold_value is not None:
                    all_data[current_cfg][step_name] = gold_value
                row += 1
            elif cell_value == '合计':
                row += 2  # 跳过合计行和空行
            else:
                row += 1
        
        return all_data
    except Exception as e:
        print(f"读取现有报表失败: {e}")
        return {}

def merge_report_data(existing_data, new_data):
    """合并报表数据，用新数据覆盖旧数据"""
    merged_data = existing_data.copy()
    
    for cfg_name, step_data in new_data.items():
        merged_data[cfg_name] = step_data
    
    return merged_data

def generate_gold_report(all_data, device_name, date_str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "金币统计报表"

    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    total_fill = PatternFill(start_color='e8f6f3', end_color='e8f6f3', fill_type='solid')
    summary_fill = PatternFill(start_color='e74c3c', end_color='e74c3c', fill_type='solid')
    summary_font = Font(bold=True, color='FFFFFF', size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    border = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

    ws.append(['📊 金币统计报表'])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = center_align
    ws.append(['设备', device_name])
    ws.append(['日期', date_str])
    ws.append([])

    total_all_gold = 0
    current_row = 5

    for cfg_name, step_data in all_data.items():
        cfg_total_gold = sum(step_data.values())
        total_all_gold += cfg_total_gold

        ws.cell(row=current_row, column=1, value=cfg_name)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:B{current_row}')
        current_row += 1

        ws.cell(row=current_row, column=1, value='STEP名称')
        ws.cell(row=current_row, column=1).fill = header_fill
        ws.cell(row=current_row, column=1).font = header_font
        ws.cell(row=current_row, column=2, value='金币数')
        ws.cell(row=current_row, column=2).fill = header_fill
        ws.cell(row=current_row, column=2).font = header_font
        for col in ['A', 'B']:
            ws.cell(row=current_row, column=1 if col == 'A' else 2).alignment = center_align
            ws.cell(row=current_row, column=1 if col == 'A' else 2).border = border
        current_row += 1

        for step_name, gold in step_data.items():
            ws.cell(row=current_row, column=1, value=step_name)
            ws.cell(row=current_row, column=1).alignment = left_align
            ws.cell(row=current_row, column=1).border = border
            ws.cell(row=current_row, column=2, value=gold)
            ws.cell(row=current_row, column=2).alignment = center_align
            ws.cell(row=current_row, column=2).border = border
            current_row += 1

        ws.cell(row=current_row, column=1, value='合计')
        ws.cell(row=current_row, column=1).fill = total_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True, color='27ae60')
        ws.cell(row=current_row, column=2, value=cfg_total_gold)
        ws.cell(row=current_row, column=2).fill = total_fill
        ws.cell(row=current_row, column=2).font = Font(bold=True, color='27ae60')
        for col in ['A', 'B']:
            ws.cell(row=current_row, column=1 if col == 'A' else 2).alignment = center_align
            ws.cell(row=current_row, column=1 if col == 'A' else 2).border = border
        current_row += 2

    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws.cell(row=current_row, column=1, value='总金币数')
    ws.cell(row=current_row, column=1).fill = summary_fill
    ws.cell(row=current_row, column=1).font = summary_font
    ws.cell(row=current_row, column=1).alignment = center_align
    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws.cell(row=current_row, column=1, value=total_all_gold)
    ws.cell(row=current_row, column=1).fill = summary_fill
    ws.cell(row=current_row, column=1).alignment = center_align
    ws.cell(row=current_row, column=1).font = Font(bold=True, color='FFFFFF', size=20)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer


def process(check):
    # jinBiPos = textFind('立即领取')
    # rect = jinBiPos.rect
    # print('jinBiPos', jinBiPos, 'rect', rect)
    # print(rect.width() + 100, rect.height() + 100)
    #
    # rectTmp = [rect.left - 100, rect.top - 300, rect.left + rect.width(), rect.top + rect.height() + 100]
    # res = Ocr().paddleocr_v3()
    # if res:
    #     for r in res:
    #         # print(r)
    #         print(r.text)  # 打印出文本
    #         # print(r['rect'])  # 范围
    #         # print(r['center_x'], r['center_y'])  # 识别范围
    #         # print(r['confidence'])  # 可信度
    #     print('end')
    # while True:
    #     swipeUp()
    #     time.sleep(randint(60, 100))
    Ocr.set_engine("mlkit")
    print(f'process {check}')
    
    device_name = Device.name()
    
    for _ in range(100):
        today_date = datetime.now().strftime('%Y%m%d')
        run_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        all_cfg_data = {}
        
        # 运行每个cfg并收集数据
        for cfg in procList:
            if cfg in check:
                step_data = procSingle(cfg, check[cfg], run_timestamp, today_date)
                all_cfg_data[cfg] = step_data
        
        # 所有cfg运行完一次后，生成并上传每次执行的报表（保留原逻辑）
        if all_cfg_data:
            # 上传每次执行的报表到 log 目录，结构：log/{device}/{date}/{timestamp}
            excel_buffer = generate_gold_report(all_cfg_data, device_name, run_timestamp)
            folder_path = f"log/{device_name}/{today_date}/{run_timestamp}"
            filename = "金币统计报表.xlsx"
            logger.upload_excel_report(excel_buffer, filename, folder_path)
            
            # 更新当天的汇总报表到 report 目录
            report_folder_path = f"report/{device_name}/{today_date}"
            report_filename = "金币统计报表.xlsx"
            
            # 下载现有的当天汇总报表
            existing_excel_buffer = logger.download_excel_report(report_filename, report_folder_path)
            existing_data = read_existing_report_data(existing_excel_buffer)
            
            # 合并新数据到现有数据
            merged_data = merge_report_data(existing_data, all_cfg_data)
            
            # 生成并上传更新后的汇总报表
            merged_excel_buffer = generate_gold_report(merged_data, device_name, today_date)
            logger.upload_excel_report(merged_excel_buffer, report_filename, report_folder_path)